# -*- coding:utf-8 -*-

import config
import dump
import os
import openpyxl

import sys
reload(sys)

sys.setdefaultencoding("utf-8")

ROOT_PATH = r"C:\Program Files (x86)\CCBComponents"
STUB_PATH = r"stub_2_0.txt"
FIXME_PATH = r"wine_2_0_fixme.txt"


def getDllsRecursive(path):
    dll_list = list()
    files = os.listdir(path)
    for file_ in files:
        if os.path.isdir(os.path.join(path, file_)):
            dll_list.extend(getDllsRecursive(os.path.join(path, file_)))
        elif file_.lower().endswith(".dll") or file_.lower().endswith(".exe") or file_.lower().endswith(".sys"):
            dll_list.append((file_, path))
    return dll_list


def main():
    dll_list = getDllsRecursive(ROOT_PATH)

    f = open("res2.txt", "w")
    for dll in dll_list:
        dll_name = dll[0]
        dll_path = dll[1]

        load, delay_load = dump.imports(os.path.join(dll_path, dll_name))

        path = os.path.relpath(dll_path, ROOT_PATH)

        for l in load:
            name = l["name"]
            for api in l["api"]:

                try:
                    f.write(dll_name + "\t")
                    f.write(path.encode("gbk"))
                    f.write("\t" + name + "\t" + api + "\t" + "load" + "\n")
                except:
                    print path
                    f.write("\n")
                    pass

        for l in delay_load:
            name = l["name"]
            for api in l["api"]:
                try:
                    f.write(dll_name + "\t")
                    f.write(path.encode("gbk"))
                    f.write("\t" + name + "\t" + api + "\t" + "delayload" + "\n")
                except:
                    print path
                    f.write("\n")
                    pass

    f.close()


def wineDLLsSet(path):
    dll_set = set()
    for line in open(FIXME_PATH):
        if ":" in line:
            line = line.replace(":","").replace("\n","")
            dll_set.add(line.lower())
    return  dll_set


def isDLLsStub(dll,api):
    isStub = "false"
    dll = dll.lower()
    for line in open(STUB_PATH):
        if ":" in line:
            dllName = line.replace("\n","").replace(":","").lower()
            # print dllName
        else:
            apiName = line.replace("\n","")
            # print apiName
            if dll == dllName and api == apiName :
                isStub = "true"
    return isStub


def isDLLsFixme(dll,api):
    isFixme = "false"
    dll = dll.lower()
    for line in open(FIXME_PATH):
        if ":" in line:
            dllName = line.replace(":","").replace("\n","")
            # print dllName
        else:
            apiName = line.replace(" ","").replace("\n","")
            # print apiName
            if dll == dllName and api == apiName:
                isFixme = "true"
    return isFixme


def getDLL_APIStatus(dll, api, status_dict, dll_set):
    if dll.lower().endswith(".dll"):
        dll = dll.lower().replace(".dll", "")
    if dll.lower() not in dll_set:
        return "DllNotInWine"
    if (dll, api) in status_dict:
        return status_dict[(dll, api)]
    else:
        if isDLLsStub(dll, api) == 'true':
            status_dict[(dll, api)] = "stub"
        elif isDLLsFixme(dll, api) == 'true':
            status_dict[(dll, api)] = "fixme"
        else:
            status_dict[(dll, api)] = ""
    return status_dict[(dll, api)]



def main2():
    app_list = list()
    print os.listdir(ROOT_PATH)
    for file in os.listdir(ROOT_PATH):
        if os.path.isdir(os.path.join(ROOT_PATH, file)):
            print file
            app_list.append(file)
    print app_list

    dll_api_status_dict = dict()
    dll_set = wineDLLsSet(FIXME_PATH)

    for app in app_list:
        installer_list = list()
        expand_list = list()

        for file in os.listdir(os.path.join(ROOT_PATH, app)):
            if os.path.isdir(os.path.join(ROOT_PATH, app, file)):
                expand_list.append(file)
            elif file.endswith(".exe") or file.endswith(".msi"):
                installer_list.append(file)
        """
        for installer in installer_list:
            f = open("./installer/" + app+"_"+installer+".txt", "w")

            load, delay_load = dump.imports(os.path.join(ROOT_PATH, app, installer))

            path = os.path.relpath(installer)

            for l in load:
                name = l["name"]
                for api in l["api"]:

                    try:
                        f.write(installer + "\t")
                        f.write(path.encode("gbk"))
                        f.write("\t" + name + "\t" + api + "\t" + "load" + "\n")
                    except:
                        print path
                        f.write("\n")
                        pass

            for l in delay_load:
                name = l["name"]
                for api in l["api"]:
                    try:
                        f.write(installer + "\t")
                        f.write(path.encode("gbk"))
                        f.write("\t" + name + "\t" + api + "\t" + "delayload" + "\n")
                    except:
                        print path
                        f.write("\n")
                        pass
            f.close()

        for expand in expand_list:
            f = open("./expand/" + app+"_"+expand+".txt", "w")

            dll_list = getDllsRecursive(os.path.join(ROOT_PATH, app, expand))
            for dll in dll_list:
                dll_name = dll[0]
                dll_path = dll[1]

                load, delay_load = dump.imports(os.path.join(dll_path, dll_name))

                path = os.path.relpath(dll_path, ROOT_PATH)

                for l in load:
                    name = l["name"]
                    for api in l["api"]:

                        try:
                            f.write(dll_name + "\t")
                            f.write(path.encode("gbk"))
                            f.write("\t" + name + "\t" + api + "\t" + "load" + "\n")
                        except:
                            print path
                            f.write("\n")
                            pass

                for l in delay_load:
                    name = l["name"]
                    for api in l["api"]:
                        try:
                            f.write(dll_name + "\t")
                            f.write(path.encode("gbk"))
                            f.write("\t" + name + "\t" + api + "\t" + "delayload" + "\n")
                        except:
                            print path
                            f.write("\n")
                            pass
            f.close()
        """


        for installer in installer_list:
            workbook = openpyxl.Workbook()
            booksheet = workbook.get_active_sheet()
            booksheet2 = workbook.create_sheet()

            dll_api_set = set()

            load, delay_load = dump.imports(os.path.join(ROOT_PATH, app, installer))

            path = os.path.relpath(installer)

            index = 1
            for l in load:
                name = l["name"]
                if len(l["api"]) == 0:
                    booksheet.cell(row=index, column=1, value="API_NOT_FOUND")
                    index += 1
                for api in l["api"]:
                    booksheet.cell(row=index, column=1, value=installer)
                    booksheet.cell(row=index, column=2, value=path)
                    booksheet.cell(row=index, column=3, value=name)
                    booksheet.cell(row=index, column=4, value=api)
                    booksheet.cell(row=index, column=5, value="load")
                    booksheet.cell(row=index, column=6, value=getDLL_APIStatus(name, api, dll_api_status_dict, dll_set))
                    index += 1
                    dll_api_set.add((name, api, getDLL_APIStatus(name, api, dll_api_status_dict, dll_set)))

            for l in delay_load:
                name = l["name"]
                if len(l["api"]) == 0:
                    booksheet.cell(row=index, column=1, value="API_NOT_FOUND")
                    index += 1
                for api in l["api"]:
                    booksheet.cell(row=index, column=1, value=installer)
                    booksheet.cell(row=index, column=2, value=path)
                    booksheet.cell(row=index, column=3, value=name)
                    booksheet.cell(row=index, column=4, value=api)
                    booksheet.cell(row=index, column=5, value="delayload")
                    booksheet.cell(row=index, column=6, value=getDLL_APIStatus(name, api, dll_api_status_dict, dll_set))
                    index += 1
                    dll_api_set.add((name, api, getDLL_APIStatus(name, api, dll_api_status_dict, dll_set)))

            dll_api_list = list(dll_api_set)
            dll_api_list.sort()

            index = 1
            for d_a in dll_api_list:
                booksheet2.cell(row=index, column=1, value=d_a[0])
                booksheet2.cell(row=index, column=2, value=d_a[1])
                booksheet2.cell(row=index, column=3, value=d_a[2])
                index += 1

            workbook.save("./installer/" + app + "_" + installer + ".xlsx")


        for expand in expand_list:
            workbook = openpyxl.Workbook()
            booksheet = workbook.get_active_sheet()
            booksheet2 = workbook.create_sheet()

            dll_api_set = set()

            index = 1
            dll_list = getDllsRecursive(os.path.join(ROOT_PATH, app, expand))

            dll_lower_list = map(lambda x: x[0].lower(), dll_list)

            for dll in dll_list:
                dll_name = dll[0]
                dll_path = dll[1]

                load, delay_load = dump.imports(os.path.join(dll_path, dll_name))

                path = os.path.relpath(dll_path, ROOT_PATH)

                for l in load:
                    name = l["name"]

                    if (name).lower().encode("utf-8") in dll_lower_list:
                        continue

                    if len(l["api"]) == 0:
                        booksheet.cell(row=index, column=1, value="API_NOT_FOUND")
                        index += 1

                    for api in l["api"]:
                        booksheet.cell(row=index, column=1, value=dll_name)
                        booksheet.cell(row=index, column=2, value=path)
                        booksheet.cell(row=index, column=3, value=name)
                        booksheet.cell(row=index, column=4, value=api)
                        booksheet.cell(row=index, column=5, value="load")
                        booksheet.cell(row=index, column=6, value=getDLL_APIStatus(name, api, dll_api_status_dict, dll_set))

                        dll_api_set.add((name, api, getDLL_APIStatus(name, api, dll_api_status_dict, dll_set)))

                        index += 1


                for l in delay_load:
                    name = l["name"]

                    if (name).lower().encode("utf-8") in dll_lower_list:
                        continue

                    if len(l["api"]) == 0:
                        booksheet.cell(row=index, column=1, value="API_NOT_FOUND")
                        index += 1

                    for api in l["api"]:
                        booksheet.cell(row=index, column=1, value=dll_name)
                        booksheet.cell(row=index, column=2, value=path)
                        booksheet.cell(row=index, column=3, value=name)
                        booksheet.cell(row=index, column=4, value=api)
                        booksheet.cell(row=index, column=5, value="delayload")
                        booksheet.cell(row=index, column=6, value=getDLL_APIStatus(name, api, dll_api_status_dict, dll_set))

                        dll_api_set.add((name, api, getDLL_APIStatus(name, api, dll_api_status_dict, dll_set)))

                        index += 1

            dll_api_list = list(dll_api_set)
            dll_api_list.sort()

            index = 1
            for d_a in dll_api_list:
                booksheet2.cell(row=index, column=1, value=d_a[0])
                booksheet2.cell(row=index, column=2, value=d_a[1])
                booksheet2.cell(row=index, column=3, value=d_a[2])
                index += 1

            workbook.save("./expand/" + app+"_"+expand+".xlsx")

def dump_app(root_path):
    pass



if __name__ == "__main__":
    main2()
