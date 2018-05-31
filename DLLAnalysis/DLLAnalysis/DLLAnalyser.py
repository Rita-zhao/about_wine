# -*- coding: utf-8 -*-
from models import DLL, API
from xml.dom import minidom
from nirsoftSpider import analysis_index as windows_analysis_index
from wineAPISpider import analysis_index as wine_analysis_index
import networkx as nx
import matplotlib.pyplot as plt
import xlwt
import openpyxl
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from tools import analyse_spec

class DLLAnalyser:
    def __init__(self):
        self.dlls = dict()

    def load_from_xml(self, path):
        dom = minidom.parse(path)
        root = dom.documentElement
        dll_node_list = root.getElementsByTagName("dll")
        for dll_node in dll_node_list:
            dll = DLL()
            dll.xml_to_dll(dll_node)
            self.dlls[dll.name] = dll

        self.__generate_dependence()


    def write_to_xml(self, path):
        dll_name_list = []
        for dll_name in self.dlls.keys():
            dll_name_list.append(dll_name)
        dll_name_list.sort()

        impl = minidom.getDOMImplementation()
        dom = impl.createDocument(None, "dll_list", None)
        root = dom.documentElement

        for dll_name in dll_name_list:
            dll_node = self.dlls[dll_name].dll_to_xml(dom)
            root.appendChild(dll_node)

        f = open(path, "w")
        dom.writexml(f, "\t", "\t", "\n", "iso-8859-1")

    def load_from_html(self, root_path, type):
        if type.lower() == "wine":
            self.dlls = wine_analysis_index(root_path)
        elif type.lower() == "windows":
            self.dlls = windows_analysis_index(root_path)
        self.__generate_dependence()

    def add_info_from_spec(self, path):
        dll_spec_list = analyse_spec(path)
        for dll_spec in dll_spec_list:
            dll_name = dll_spec[0].lower()
            spec_list = dll_spec[1]
            dll = None
            dll_exist = False
#            dll_full_name = dll_name
#判断 spec文件中读出的dll名字是否已经存在于从html获取的dll——list中，当名字中没有.时，后缀加dll，当有.时，名字保持不变
            for dll in self.dlls.keys():
                dot_index = dll_name.find(".")
                if dot_index == -1:
                    dll_full_name = dll_name + ".dll"
                else:
                    dll_full_name = dll_name
                if dll == dll_full_name:
                    dll_exist = True
                    dll = self.dlls[dll_full_name]
                    break
#如果已经存在于dll_list中，则判断api是否存在于api_list；
#如果存在，将forward信息补充上即可，同时，state信息进行更新；
#如果不存在，补充api的完整信息，但spec中无法获取参数信息，即从spec中补充的api信息是没有参数信息的
            if dll_exist:
                api_list = dll.api_list
                for spec_line in spec_list:
                    (state, api_name, dll_name, forward_api) = self.__spec_line(spec_line)
                    for api in dll.api_list:
                        if api.name == api_name:
                            if spec_line[3].find("(") != -1 and (api.synopsis == None or api.synopsis == ""):
                                api.synopsis = spec_line[3]
                            api.state = "from spec: " + spec_line[1] + spec_line[2]
                            if spec_line[3]:
                                api.state = api.state + spec_line[3]                                
                            api.forward_api = forward_api
                            break
                    else:
                        api = API()
                        api.name = api_name
                        api.state = "from spec: " + spec_line[1] + spec_line[2]
                        if spec_line[3]:
                            api.state = api.state + spec_line[3]
                        api.forward_api = forward_api
                        dll.api_list.append(api)
#
            else:
                dll = DLL() 
                dot_index_ = dll_name.find(".")
                if dot_index_ == -1:
                    dll_full_name = dll_name + ".dll"
                else:
                    dll_full_name = dll_name    
                dll.name = dll_full_name

                for spec_line in spec_list:
                    (state, api_name, dll_name, forward_api) = self.__spec_line(spec_line)
                    api = API()
                    api.name = api_name
                    api.state = "from spec: " + spec_line[1] + spec_line[2]
                    if spec_line[3]:
                        api.state = api.state + spec_line[3]
                    api.forward_api = forward_api
                    dll.api_list.append(api)
                self.dlls[dll.name] = dll


    def __spec_line(self, spec_line):
        state = spec_line[1]
        api_synopsis = spec_line[3]
        api_impl = spec_line[4]

        index = api_synopsis.find("(")
        if index != -1:
            api_name = api_synopsis[:index].strip()
        else:
            api_name = api_synopsis.strip()

        dll = None
        forward_api = api_impl
        '''
        if api_impl != None:
            index = api_impl.find(".")
            if index != -1:
                dll = api_impl[:index]
                forward_api = api_impl[index+1:]
            else:
                forward_api = api_impl
        '''
        return (state, api_name, dll, forward_api)

    def generate_gv_dll_relation(self, path):
        f = open(path, "w")
        f.write("digraph G {\n")
        for dll in self.dlls.values():
            for depend_dll in dll.dependent_dll_list:
                f.write("\t" + dll.name.lower().replace(".", "_") + " -> " + depend_dll.lower().replace(".", "_") + ";\n")
        f.write("}\n")
        f.close()

    def print_graph(self):
        G = nx.Graph()
        for dll in self.dlls.values():
            if dll.dependent_dll_list == None or len(dll.dependent_dll_list) == 0:
                G.add_node(dll.name.lower())
            else:
                for depend_dll in dll.dependent_dll_list:
                    G.add_edge(dll.name.lower(), depend_dll.lower())
        nx.draw(G)
        plt.show()
    def __generate_dependence(self):
        for dll in self.dlls.values():
            for depend_dll in dll.dependent_dll_list:
                if depend_dll.lower() in self.dlls:
                    self.dlls[depend_dll.lower()].call_dll_list.append(dll.name.lower())
                else:
                    print depend_dll.lower() + " is not a dll"
            dll.depend_dll_num = len(dll.dependent_dll_list)
            dll.call_dll_num = len(dll.call_dll_list)

    def __write_xls(self, path):
        # not used
        workbook = xlwt.Workbook(encoding='utf-8')
        booksheet = workbook.add_sheet("DLL", cell_overwrite_ok=True)

        dll_list = self.dlls.values()
        dll_list.sort(key=lambda x: x.name)

        index = 0
        booksheet.write(index, 0, "dll.name")
        booksheet.write(index, 1, "dll.intro")
        booksheet.write(index, 2, "dll.depend_dll_num: how many dlls does this dll depend on")
        booksheet.write(index, 3, "dll.call_dll_num: how mand dlls call this dll")
        index += 1

        for dll in dll_list:
            booksheet.write(index, 0, dll.name)
            booksheet.write(index, 1, dll.intro)
            booksheet.write(index, 2, dll.depend_dll_num)
            booksheet.write(index, 3, dll.call_dll_num)
            index += 1

        booksheet2 = workbook.add_sheet("Dependence", cell_overwrite_ok=True)

        index = 0
        booksheet2.write(index, 0, "dll.name")
        booksheet2.write(index, 1, "dependent_dll")
        index += 1

        for dll in dll_list:
            dependent_dll_list = dll.dependent_dll_list
            dependent_dll_list.sort()
            for dependent_dll in dependent_dll_list:
                booksheet2.write(index, 0, dll.name)
                booksheet2.write(index, 1, dependent_dll)
                index += 1

        workbook.save(path)

    def write_xlsx(self, path):
        workbook = openpyxl.Workbook()
        booksheet = workbook.create_sheet("DLL")

        dll_list = self.dlls.values()
        dll_list.sort(key=lambda x: x.name)

        booksheet.cell(row=1, column=1, value="dll.name")
        booksheet.cell(row=1, column=2, value="dll.intro")
        booksheet.cell(row=1, column=3, value="dll.depend_dll_num")
        booksheet.cell(row=1, column=4, value="dll.call_dll_num")

        index = 2

        for dll in dll_list:
            booksheet.cell(row=index, column=1, value=dll.name)
            booksheet.cell(row=index, column=2, value=dll.intro)
            booksheet.cell(row=index, column=3, value=dll.depend_dll_num)
            booksheet.cell(row=index, column=4, value=dll.call_dll_num)
            index += 1

        booksheet2 = workbook.create_sheet("Dependence")

        booksheet2.cell(row=1, column=1, value="dll.name")
        booksheet2.cell(row=1, column=2, value="dll.dependent_dll")

        index = 2
        for dll in dll_list:
            dependent_dll_list = dll.dependent_dll_list
            dependent_dll_list.sort()
            for dependent_dll in dependent_dll_list:
                booksheet2.cell(row=index, column=1, value=dll.name)
                booksheet2.cell(row=index, column=2, value=dependent_dll)

                if dependent_dll not in self.dlls:
                    booksheet2.cell(row=index, column=2).font = Font(color=colors.RED)

                index += 1

        workbook.save(path)


if __name__ == "__main__":
    dll_analyser = DLLAnalyser()
    dll_analyser.load_from_html(r"D:\winXP_dlls/", "windows")
    dll_analyser.write_xlsx(r"D:\winXP_dlls.xlsx")

    dll_analyser = DLLAnalyser()
    dll_analyser.load_from_html(r"D:\win7_dlls/", "windows")
    dll_analyser.write_xlsx(r"D:\win7_dlls.xlsx")

    dll_analyser = DLLAnalyser()
    dll_analyser.load_from_html(r"D:\win8_dlls/", "windows")
    dll_analyser.write_xlsx(r"D:\win8_dlls.xlsx")

    dll_analyser = DLLAnalyser()
    dll_analyser.load_from_html(r"D:\win10_dlls/", "windows")
    dll_analyser.write_xlsx(r"D:\win10_dlls.xlsx")