#coding:utf-8

import os
from urllib2 import urlopen
import openpyxl
import re


def writeFile(rootdir, file_path, content, do_write_file=True):
    if do_write_file:
        if not os.path.exists(rootdir):
            os.makedirs(rootdir)
        f = open(rootdir + "/" + file_path, "w")
        f.write(content)
        f.close()


def readHtml(path):
    try:
        if path.startswith("http") or path.startswith("https"):
            page = urlopen(path)
            html = page.read()
        else:
            page = open(path)
            html = page.read()
            page.close()
        return html.decode("ISO-8859-1").encode("utf-8")
    except Exception as e:
        # print e, path
        return "ERROR!" + e.__str__()


def generate_xlsx_for_doxygen(path, out_path):
    dll_dir_list = os.listdir(path)

    workbook = openpyxl.Workbook()
    booksheet = workbook.get_active_sheet()

    index = 1
    booksheet.cell(row=index, column=1, value=u"层次")
    booksheet.cell(row=index, column=2, value=u"模板(Name,Type,generateBuild)/子模板(Name,Type,generateBuild)/...")
    booksheet.cell(row=index, column=3, value="Path or Src")

    index = 2
    for dll_dir in dll_dir_list:
        file_list = os.listdir(os.path.join(path, dll_dir))
        booksheet.cell(row=index, column=1, value="DLL")
        booksheet.cell(row=index, column=2, value=dll_dir)
        booksheet.cell(row=index, column=3, value="NULL")
        booksheet.cell(row=index, column=4, value="1")
        file_name_list = []
        for f in file_list:
            if os.path.isfile(os.path.join(path, dll_dir, f)):
                print dll_dir, f
                file_name_list.append(os.path.join("dlls", dll_dir, f))

        booksheet.cell(row=index, column=5, value=",".join(file_name_list))
        index += 1

    workbook.save(out_path)


def analyse_spec(path):
    dll_dir_list = os.listdir(path)
#    state_set = set()

    result = []

    for dll_dir in dll_dir_list:
        file_list = os.listdir(os.path.join(path, dll_dir))
        for f in file_list:
            if f.endswith(".spec"):

                spec_file = open(os.path.join(path, dll_dir, f))
                spec_list = []
                for line in spec_file:
                    if line.strip() == "":
                        continue
                    if line.strip().startswith("#"):
                        '''
                        if line.startswith(" #@"):
                            line.strip(" #")
                            print line
                        if line.startswith(" # @"):
                            line.strip(" # ")     
                            print line
                        else:
                            continue
                        '''
                        with open("C:/Users/iscas/Desktop/Wingear/dll&api/log_comment", "a") as x:
                            x.write(line+"\n")                        
                        continue
                    res_list = __analyse_spec_line(line.strip())
                    spec_list.append(res_list)
#                    state_set.add(res_list[1])
                spec_file.close()
                result.append((dll_dir, spec_list))
                break

#    print state_set
    return result


def __analyse_spec_line(line):
    s = r"(\d+|@)\s+(\S+)\s+((?:-\S+\s+)*)((?:[^\s\(\)])+(?:\s*\(.*\))?)(?:\s+(\S+))?\s*(#.*)?"
    # num|@   stub|stdcall...   -xxx=xxx   funcname(param)   otherfuncname  #comment
    pattern = re.compile(s)
    match = pattern.match(line)
    if match:
        return match.groups()
    else:
        print "error:", line
        return []
